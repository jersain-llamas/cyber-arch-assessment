# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import RadarChart, Reference
from openpyxl.utils import get_column_letter
import loader as _L
from loader import PILLARS, SUB, ITEMS, ORDER, SCALE

NAVY="1F3A5F"; GOLD="C8A24B"; LIGHT="EAF0F6"; GREY="6B7785"; WHITE="FFFFFF"; INPUT="FFF6CC"
FONT="Arial"
def F(sz=10,b=False,c="000000"): return Font(name=FONT,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="C9D2DC")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)
tl=Alignment(horizontal="left",vertical="top",wrap_text=True)
ctop=Alignment(horizontal="center",vertical="top",wrap_text=True)

wb=Workbook()

# ---------------- INICIO ----------------
ws=wb.active; ws.title="Inicio"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=112
def ln(r,t,sz=10,b=False,c="000000",bg=None,h=None):
    cl=ws.cell(row=r,column=2,value=t); cl.font=F(sz,b,c); cl.alignment=tl
    if bg: cl.fill=fill(bg)
    if h: ws.row_dimensions[r].height=h
ws.row_dimensions[1].height=8
ln(2,"AUTOEVALUACIÓN DE ARQUITECTURA LEGAL DE CIBERSEGURIDAD",18,True,NAVY)
ln(3,"Arquetipo legislativo mínimo viable · América Latina y el Caribe · 10 pilares / 59 subpilares / 177 ítems",11,False,GREY)
ws.row_dimensions[4].height=6
ln(5,"  CÓMO LLENARLO (3 pasos)",12,True,WHITE,NAVY,h=26)
ln(6,"1.  Abra la hoja «Autoevaluación». Cada subpilar tiene 3 ítems. En la columna amarilla «Nivel (0-4)»",10)
ln(7,"     elija un número de 0 a 4 (hay menú desplegable al hacer clic en la celda). Esa es la ÚNICA columna obligatoria.",10)
ln(8,"2.  Opcional: anote en «Evidencia» el artículo de su ley y en «Nota» qué falta para subir de nivel.",10)
ln(9,"3.  Las hojas «Resumen» y «Tablero» se calculan solas: promedian los ítems por subpilar y por pilar.",10)
ws.row_dimensions[10].height=6
ln(11,"  SI LA GRÁFICA O LOS NÚMEROS NO CAMBIAN",12,True,WHITE,"B45309",h=26)
ln(12,"•  Use Excel de escritorio o LibreOffice. En Google Sheets el radar puede importarse como imagen fija;",10)
ln(13,"   las celdas de color de las hojas «Resumen» y «Tablero» sí se actualizan en cualquier programa.",10)
ln(14,"•  Si nada recalcula: menú Fórmulas → Opciones de cálculo → Automático (o pulse F9 / Cmd + =).",10)
ln(15,"•  El radar promedia muchos ítems: cambiar un solo ítem lo mueve poco. Mire mejor la hoja «Resumen».",10)
ws.row_dimensions[16].height=6
ln(17,"  ESCALA 0-4 Y LAS DOS LECTURAS",12,True,WHITE,NAVY,h=26)
ln(18,"Niveles 0-3 = arquitectura legal (alcance del documento). Nivel 4 = capa operativa verificable (extensión). Detalle en «Escala».",10)
ln(19,"•  Cobertura legal = % de subpilares con promedio ≥ 3.   •  Madurez operativa = % de subpilares con promedio = 4.",10)
ws.row_dimensions[20].height=6
ln(21,"  NOTAS",12,True,WHITE,NAVY,h=26)
ln(22,"•  Tesis del marco: el problema regional es de ARQUITECTURA JURÍDICA COHERENTE, no de mera ausencia de normas.",10)
ln(23,"•  El «Tier» (Núcleo/Intermedio/Avanzado) es una secuencia de adopción PROPUESTA y editable, no del documento original.",10)
ln(24,"•  Las referencias (NIS2, DORA, CRA, CER, CSA, Chile 21.663…) son anclas de redacción, no de copia literal.",10,c=GREY)
for r in range(5,25):
    ws.cell(row=r,column=2).border=Border(left=thin,right=thin)
ws.cell(row=25,column=2).border=Border(top=thin)

# ---------------- ESCALA ----------------
es=wb.create_sheet("Escala"); es.sheet_view.showGridLines=False
es.column_dimensions["A"].width=8; es.column_dimensions["B"].width=30; es.column_dimensions["C"].width=95
es.cell(row=1,column=1,value="RÚBRICA DE MADUREZ (0-4)").font=F(14,True,NAVY)
for j,h in enumerate(["Nivel","Etiqueta","Descripción"],1):
    c=es.cell(row=3,column=j,value=h); c.font=F(10,True,WHITE); c.fill=fill(NAVY); c.alignment=center; c.border=border
sc=["F4CCCC","FCE5CD","FFF2CC","D9EAD3","B6D7A8"]
for i,(n,lab,desc) in enumerate(SCALE):
    r=4+i
    es.cell(row=r,column=1,value=n).alignment=center
    es.cell(row=r,column=2,value=lab).alignment=left
    es.cell(row=r,column=3,value=desc).alignment=left
    for j in range(1,4):
        cell=es.cell(row=r,column=j); cell.border=border; cell.font=F(10); cell.fill=fill(sc[i])
    es.row_dimensions[r].height=34
es.cell(row=10,column=1,value="Lectura: nivel ≥ 3 = ya está en una ley especial (cobertura legal). Nivel 4 = además operativo y verificable.").font=F(9,False,GREY)

# ---------------- AUTOEVALUACIÓN ----------------
au=wb.create_sheet("Autoevaluación"); au.sheet_view.showGridLines=False
cols=[("Pilar",6),("Ref",6),("Subpilar",26),("Tier",11),("Ítem a verificar",70),
      ("Nivel\n(0-4)",8),("Evidencia (art./sección de su ley)",32),("Nota / acción",32)]
for j,(h,w) in enumerate(cols,1):
    au.column_dimensions[get_column_letter(j)].width=w
    c=au.cell(row=1,column=j,value=h); c.font=F(10,True,WHITE); c.fill=fill(NAVY); c.alignment=center; c.border=border
au.row_dimensions[1].height=30
tfill={"Núcleo":fill("FCE8D5"),"Intermedio":fill("E7EEF6"),"Avanzado":fill("EFEFEF")}
band=["FFFFFF","F4F7FA"]
r=2; first=2
for bi,ref in enumerate(ORDER):
    L=ref[0]; sub,tier,src=SUB[ref]; items=ITEMS[ref]
    bg=band[bi%2]
    for k,it in enumerate(items):
        au.cell(row=r,column=1,value=(L if k==0 else None)).alignment=ctop
        au.cell(row=r,column=2,value=ref).alignment=ctop
        au.cell(row=r,column=3,value=(sub if k==0 else None)).alignment=tl
        tc=au.cell(row=r,column=4,value=(tier if k==0 else None)); tc.alignment=ctop
        if k==0: tc.fill=tfill[tier]; tc.font=F(9,True,GREY)
        au.cell(row=r,column=5,value=f"• {it}").alignment=tl
        nv=au.cell(row=r,column=6,value=0); nv.alignment=center; nv.font=F(11,True); nv.fill=fill(INPUT)
        au.cell(row=r,column=7,value="").alignment=tl
        au.cell(row=r,column=8,value="").alignment=tl
        for j in range(1,9):
            cell=au.cell(row=r,column=j); cell.border=border
            if cell.font.size is None: cell.font=F(9)
            if j in (3,5,7,8) and cell.fill.fgColor.rgb in (None,"00000000"): cell.fill=fill(bg)
        au.cell(row=r,column=7).fill=fill(INPUT); au.cell(row=r,column=8).fill=fill(INPUT)
        au.row_dimensions[r].height=30
        r+=1
last=r-1
dv=DataValidation(type="list",formula1='"0,1,2,3,4"',allow_blank=False)
dv.error="Use un nivel de 0 a 4 (ver hoja Escala)."; dv.errorTitle="Nivel inválido"
au.add_data_validation(dv); dv.add(f"F{first}:F{last}")
au.conditional_formatting.add(f"F{first}:F{last}",
    ColorScaleRule(start_type="num",start_value=0,start_color="F4737A",
                   mid_type="num",mid_value=2,mid_color="FFE08A",
                   end_type="num",end_value=4,end_color="7BC07F"))
au.freeze_panes="F2"; au.auto_filter.ref=f"A1:H{last}"
ITEM_REF=f"Autoevaluación!$B${first}:$B${last}"
ITEM_NIV=f"Autoevaluación!$F${first}:$F${last}"

# ---------------- RESUMEN SUBPILARES ----------------
rs=wb.create_sheet("Resumen"); rs.sheet_view.showGridLines=False
rcols=[("Pilar",6),("Ref",6),("Subpilar",30),("Tier",11),("Nivel\nsubpilar",10),
       ("Etiqueta",24),("¿En ley? (≥3)",13),("Prioridad",12)]
for j,(h,w) in enumerate(rcols,1):
    rs.column_dimensions[get_column_letter(j)].width=w
    c=rs.cell(row=1,column=j,value=h); c.font=F(10,True,WHITE); c.fill=fill(NAVY); c.alignment=center; c.border=border
rs.row_dimensions[1].height=30
r=2; rfirst=2
for ref in ORDER:
    L=ref[0]; sub,tier,src=SUB[ref]
    rs.cell(row=r,column=1,value=L).alignment=center
    rs.cell(row=r,column=2,value=ref).alignment=center
    rs.cell(row=r,column=3,value=sub).alignment=left
    rs.cell(row=r,column=4,value=tier).alignment=center
    rs.cell(row=r,column=5,value=f'=IFERROR(AVERAGEIF({ITEM_REF},B{r},{ITEM_NIV}),0)').number_format="0.0"
    rs.cell(row=r,column=6,value=f'=IFERROR(VLOOKUP(ROUND(E{r},0),Escala!$A$4:$B$8,2,FALSE),"")').alignment=left
    rs.cell(row=r,column=7,value=f'=IF(E{r}>=3,"Sí","No")').alignment=center
    rs.cell(row=r,column=8,value=f'=IF(E{r}>=3,"En ley",IF(D{r}="Núcleo","ALTA",IF(D{r}="Intermedio","MEDIA","BAJA")))').alignment=center
    for j in range(1,9):
        cell=rs.cell(row=r,column=j); cell.border=border
        if cell.font.size is None: cell.font=F(9)
        if j in (2,4,5,7,8): cell.alignment=center
    rs.cell(row=r,column=5).font=F(11,True)
    r+=1
rlast=r-1
rs.freeze_panes="C2"; rs.auto_filter.ref=f"A1:H{rlast}"
rs.conditional_formatting.add(f"E{rfirst}:E{rlast}",
    ColorScaleRule(start_type="num",start_value=0,start_color="F4737A",
                   mid_type="num",mid_value=2,mid_color="FFE08A",
                   end_type="num",end_value=4,end_color="7BC07F"))
rs.conditional_formatting.add(f"H{rfirst}:H{rlast}",CellIsRule(operator="equal",formula=['"ALTA"'],fill=fill("F4737A"),font=F(9,True,WHITE)))
rs.conditional_formatting.add(f"H{rfirst}:H{rlast}",CellIsRule(operator="equal",formula=['"MEDIA"'],fill=fill("FFD27A")))
rs.conditional_formatting.add(f"H{rfirst}:H{rlast}",CellIsRule(operator="equal",formula=['"En ley"'],fill=fill("D9EAD3"),font=F(9,False,"38761D")))
SREF=f"Resumen!$A${rfirst}:$A${rlast}"; SLVL=f"Resumen!$E${rfirst}:$E${rlast}"; STIER=f"Resumen!$D${rfirst}:$D${rlast}"
NS=rlast-rfirst+1

# ---------------- TABLERO ----------------
tb=wb.create_sheet("Tablero"); tb.sheet_view.showGridLines=False
for col,w in {"A":3,"B":15,"C":42,"D":13,"E":15,"F":16,"G":16}.items():
    tb.column_dimensions[col].width=w
tb.cell(row=2,column=2,value="TABLERO DE RESULTADOS").font=F(16,True,NAVY)
tb.cell(row=3,column=2,value="Se recalcula al cambiar la columna «Nivel» de la hoja Autoevaluación.").font=F(10,False,GREY)
kpis=[("Nivel promedio (0-4)",f'=IFERROR(AVERAGE({SLVL}),0)',"0.00"),
      ("Cobertura legal (≥3)",f'=COUNTIF({SLVL},">=3")/{NS}',"0%"),
      ("Madurez operativa (=4)",f'=COUNTIF({SLVL},"=4")/{NS}',"0%"),
      ("Brechas Núcleo abiertas",f'=COUNTIFS({STIER},"Núcleo",{SLVL},"<3")',"0")]
for i,(lab,fm,fmt) in enumerate(kpis):
    cL=tb.cell(row=5,column=2+i,value=lab); cL.font=F(9,True,WHITE); cL.fill=fill(NAVY); cL.alignment=center; cL.border=border
    cV=tb.cell(row=6,column=2+i,value=fm); cV.font=F(16,True,NAVY); cV.alignment=center; cV.number_format=fmt; cV.fill=fill(LIGHT); cV.border=border
    tb.row_dimensions[6].height=30
hr=9
for j,h in enumerate(["Pilar","Nombre","Subpilares","Nivel prom.","Cobertura ≥3","Madurez =4","Brechas Núcleo"]):
    c=tb.cell(row=hr,column=2+j,value=h); c.font=F(10,True,WHITE); c.fill=fill(NAVY); c.alignment=center; c.border=border
tb.row_dimensions[hr].height=24
rr=hr+1; pf=rr
for L,(name,core) in PILLARS.items():
    tb.cell(row=rr,column=2,value=L).alignment=center
    tb.cell(row=rr,column=3,value=name).alignment=left
    tb.cell(row=rr,column=4,value=f'=COUNTIF({SREF},"{L}")').alignment=center
    tb.cell(row=rr,column=5,value=f'=IFERROR(AVERAGEIF({SREF},"{L}",{SLVL}),0)').number_format="0.00"
    tb.cell(row=rr,column=6,value=f'=COUNTIFS({SREF},"{L}",{SLVL},">=3")/COUNTIF({SREF},"{L}")').number_format="0%"
    tb.cell(row=rr,column=7,value=f'=COUNTIFS({SREF},"{L}",{SLVL},"=4")/COUNTIF({SREF},"{L}")').number_format="0%"
    tb.cell(row=rr,column=8,value=f'=COUNTIFS({SREF},"{L}",{STIER},"Núcleo",{SLVL},"<3")').alignment=center
    for j in range(2,9):
        cell=tb.cell(row=rr,column=j); cell.border=border
        if cell.font.size is None: cell.font=F(9)
        if j in (4,5,6,7,8): cell.alignment=center
    rr+=1
pl=rr-1
tb.cell(row=rr,column=2,value="TOTAL").font=F(10,True,NAVY)
tb.cell(row=rr,column=4,value=f'=SUM(D{pf}:D{pl})').alignment=center
tb.cell(row=rr,column=5,value=f'=IFERROR(AVERAGE({SLVL}),0)').number_format="0.00"
tb.cell(row=rr,column=6,value=f'=COUNTIF({SLVL},">=3")/{NS}').number_format="0%"
tb.cell(row=rr,column=7,value=f'=COUNTIF({SLVL},"=4")/{NS}').number_format="0%"
tb.cell(row=rr,column=8,value=f'=COUNTIFS({STIER},"Núcleo",{SLVL},"<3")').alignment=center
for j in range(2,9):
    cell=tb.cell(row=rr,column=j); cell.border=border; cell.fill=fill(LIGHT)
    if cell.font.size is None: cell.font=F(9,True)
# color scale on pillar avg (robust feedback in any viewer)
tb.conditional_formatting.add(f"E{pf}:E{pl}",
    ColorScaleRule(start_type="num",start_value=0,start_color="F4737A",
                   mid_type="num",mid_value=2,mid_color="FFE08A",
                   end_type="num",end_value=4,end_color="7BC07F"))
# radar
ch=RadarChart(); ch.type="filled"; ch.style=2; ch.title="Nivel de madurez por pilar (0-4)"
data=Reference(tb,min_col=5,min_row=hr,max_row=pl); cats=Reference(tb,min_col=2,min_row=pf,max_row=pl)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
ch.height=10; ch.width=14; ch.y_axis.scaling.min=0; ch.y_axis.scaling.max=4
tb.add_chart(ch,"B"+str(pl+3))
tb.cell(row=pl+2,column=2,value="Radar (solo Excel/LibreOffice). Para feedback universal, mire los colores de «Nivel prom.» y de la hoja Resumen.").font=F(9,False,GREY)

_out=_L.ROOT/"dist"; _out.mkdir(parents=True,exist_ok=True)
wb.save(str(_out/"autoevaluacion_ciberseguridad_ALC.xlsx"))
pass
